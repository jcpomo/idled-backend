import io
from openpyxl import load_workbook
from pypdf import PdfReader

def extract_text(content: bytes, content_type: str) -> str:
    ct = (content_type or "").lower()
    if ct == "application/pdf":
        reader = PdfReader(io.BytesIO(content))
        pages = [(page.extract_text() or "") for page in reader.pages]
        return "\n".join(pages)
    if "spreadsheet" in ct or "excel" in ct:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts: list[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    parts.append(" ".join(cells))
        return "\n".join(parts)
    if ct.startswith("text/"):
        return content.decode("utf-8", errors="replace")
    raise ValueError(f"Tipo de documento no soportado: {content_type}")
